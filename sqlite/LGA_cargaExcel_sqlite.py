import signal, sys
import os
from openpyxl import load_workbook
import sqlite3
from sqlite3 import IntegrityError
import re


def cortarPrograma(sig, frame):
    print("\n\n[!]Saliendo ...\n")
    sys.exit(1)


# Ctlr + C
signal.signal(signal.SIGINT, cortarPrograma)

# Obtener el directorio del script
script_dir = os.path.dirname(os.path.abspath(__file__))
db_dir = os.path.join(script_dir, "db")
docs_dir = os.path.join(script_dir, "docs")

# Crear carpetas si no existen
os.makedirs(db_dir, exist_ok=True)
os.makedirs(docs_dir, exist_ok=True)

# Variables globales
fichero_permisos = os.path.join(docs_dir, "excel-cod.xlsx")

conn = sqlite3.connect(os.path.join(db_dir, "formulario.sqlite"))
cursor = conn.cursor()
wb = load_workbook(fichero_permisos)

# Función para normalizar el campo de validez

def normalizar_validez(valor):
    if not valor:
        return None, None

    v = str(valor).strip().lower()
    numeros = re.findall(r"\d+", v)

    if not numeros:
        return None, None

    num = numeros[0]

    # MESES
    if "mes" in v or "meses" in v or v.endswith("m"):
        return num, "MESES"

    # DÍAS
    if "dia" in v or "dias" in v or "días" in v:
        return num, "DÍAS"

    # SOLO NÚMERO = DÍAS
    if v.isdigit():
        return num, "DÍAS"

    return None, None

# Función para normalizar campos S/N

def normalizar_sn(valor):
    if not valor:
        return "N"  # vacío = NO

    v = str(valor).strip().lower()
    if v == "n/a":
        return "N"
    return "S"


def cargaModelos():
    cursor.execute("DELETE FROM lga_modelos")
    for nombre_hoja in wb.sheetnames:
        id_formulario = nombre_hoja
        des_modelo = ''

        if not id_formulario:
            print(f"Hoja ignorada por ID vacío: {nombre_hoja}")
            continue

        try:
            cursor.execute(
                "INSERT INTO LGA_MODELOS (ID, DES_MODELO) VALUES (?, ?)",
                (id_formulario, des_modelo)
            )
        except IntegrityError as e:
            print(f"Error insertando {id_formulario}: {e}")

    conn.commit()
    print("Todos los modelos procesados.")


def cargaPermisos():
    cursor.execute("DELETE FROM lga_permisos")
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            id_permiso = fila[10]
            des_permiso = fila[3]
            lucrativo = normalizar_sn(fila[19])
            residencia = normalizar_sn(fila[18])
            via_defecto = fila[11]
            # meses_validez = fila[7] añadire cuando se sepa cual es
            reglamento = fila[12]

            if not id_permiso or not des_permiso or not lucrativo:
                print(f"Fila {i} ignorada por campos vacíos")
                continue

            try:
                cursor.execute( #añadir meses_validez cuando se sepa cual es
                    "INSERT INTO LGA_PERMISOS (ID, DES_PERMISO, LUCRATIVO, RESIDENCIA, VIA_DEFECTO, REGLAMENTO) VALUES (?, ?, ?, ?, ?, ?)",
                    (id_permiso, des_permiso, lucrativo, residencia, via_defecto, reglamento)
                )
            except IntegrityError as e:
                print(f"Error insertando {id_permiso} en fila {i}: {e}")

        conn.commit()
        print(f"Hoja '{hoja.title}' procesada.")


def cargaViaAcceso():
    cursor.execute("DELETE FROM lga_via_acceso")

    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            id_via = fila[11]
            des_via = fila[5]

            if not id_via:
                print(f"Fila {i} ignorada por campos vacíos (ID o DES_VIA_ACCESO)")
                continue

            try:
                cursor.execute(
                    "INSERT INTO LGA_VIA_ACCESO (ID, DES_VIA_ACCESO) VALUES (?, ?)",
                    (id_via, des_via)
                )
            except IntegrityError as e:
                print(f"Error insertando {id_via} en fila {i}: {e}")

        conn.commit()
        print(f"Hoja '{hoja.title}' procesada para VIA_ACCESO.")


def cargaAutorizaciones():
    cursor.execute("DELETE FROM lga_autorizaciones")

    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        id_modelo = nombre_hoja

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            cod_meyss = fila[9]
            id_permiso = fila[10]
            id_via = fila[11]
            silencio = fila[8]

            #Campo combinado numero y meses/dias
            num_plazo, tipo_plazo = normalizar_validez(fila[7])

            if not cod_meyss or not id_permiso or not id_via or not id_modelo:
                print(f"Fila {i} ignorada por campos vacíos")
                continue

            try:
                cursor.execute(
                    "INSERT INTO LGA_AUTORIZACIONES (COD_MEYSS, ID_PERMISO, ID_VIA, ID_MODELO, NUM_PLAZO, TIPO_PLAZO, SILENCIO) VALUES (?, ?, ?, ?, ?, ?, ?)",
                    (cod_meyss, id_permiso, id_via, id_modelo, num_plazo, tipo_plazo, silencio)
                )
            except IntegrityError as e:
                print(f"Error insertando autorización en fila {i}: {e}")

        conn.commit()
        print(f"Hoja '{hoja.title}' procesada para AUTORIZACIONES.")


def menu():
    print("Seleccione una opción:")
    print("1. Cargar Modelos")
    print("2. Cargar Permisos")
    print("3. Cargar Vías de Acceso")
    print("4. Cargar Autorizaciones")
    print("0. Salir\n")


def menuExcel():
    while True:
        menu()
        opcion = input("Ingrese el número de la opción: ")
        if opcion == "1":
            cargaModelos()
        elif opcion == "2":
            cargaPermisos()
        elif opcion == "3":
            cargaViaAcceso()
        elif opcion == "4":
            cargaAutorizaciones()
        elif opcion == "0":
            print("Saliendo del programa.")
            break
        else:
            print("Opción no válida. Intente de nuevo.\n")
            continue


if __name__ == "__main__":
    menuExcel()
