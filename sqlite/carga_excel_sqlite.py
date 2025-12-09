import signal, sys
import os
from openpyxl import load_workbook
import sqlite3
from sqlite3 import IntegrityError


def cortarPrograma(sig, frame):
    print("\n\n[!]Saliendo ...\n")
    sys.exit(1)


# Ctlr + C
signal.signal(signal.SIGINT, cortarPrograma)

# Obtener el directorio del script
script_dir = os.path.dirname(os.path.abspath(__file__))
db_dir = os.path.join(script_dir, "db")
docs_dir = os.path.join(script_dir, "docs")

# Crear carpetas si no existen
os.makedirs(db_dir, exist_ok=True)
os.makedirs(docs_dir, exist_ok=True)

# Variables globales
fichero_permisos = os.path.join(docs_dir, "excel-cod.xlsx")

conn = sqlite3.connect(os.path.join(db_dir, "permisos.sqlite"))
cursor = conn.cursor()
wb = load_workbook(fichero_permisos)

def cargaExcel(tabla: str, codigo_col: int, descripcion_col: int):
    # Borrar tabla antes de cargar nuevos datos
    cursor.execute(f"DELETE FROM {tabla}")  
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        for fila in hoja.iter_rows(min_row=2, values_only=True):

            codigo = fila[codigo_col]
            descripcion = fila[descripcion_col]
            formulario = hoja.title

            if not codigo or not descripcion or not formulario:
                continue

            try:
                # INSERT dinámico usando f-string para el nombre de la tabla
                cursor.execute(
                    f"INSERT INTO {tabla} (codigo, descripcion, formulario) VALUES (?, ?, ?)",
                    (codigo, descripcion, formulario)
                )
                conn.commit()

            except IntegrityError as e:
                print(f"Error insertando {codigo} en tabla {tabla}: {e}")

def cargaAutorizacion():
    cursor.execute("DELETE FROM autorizacion")
    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            codMEYSS = fila[9]        
            codigo_permiso = fila[10]  
            codigo_via = fila[11]      
            formulario = hoja.title

            # Saltar filas incompletas
            if not codMEYSS or not codigo_permiso or not codigo_via:
                print(f"Fila {i} ignorada por valores nulos")
                continue
            try:
                cursor.execute(
                    "INSERT INTO autorizacion (codMEYSS, codigo_permiso, codigo_via, formulario) VALUES (?, ?, ?, ?)",
                    (codMEYSS, codigo_permiso, codigo_via, formulario)
                )
            except IntegrityError as e:
                # Esto capturará problemas de FK o cualquier otro error
                print(f"Error insertando {codMEYSS} en fila {i}: {e}")

        conn.commit()
        print(f"Hoja '{hoja.title}' procesada.")

def menu():
    print("Seleccione una opción:")
    print("1. Cargar Permisos")
    print("2. Cargar Vías")
    print("3. Cargar Autorizaciones")
    print("0. Salir\n")

# Menú de opciones
def menuCargaExcel():
    while True:
        menu()
        opcion = input("Ingrese el número de la opción: ")
        if opcion == "1": 
            # Cargar Permisos
            cargaExcel("permiso", 10, 3)
            continue
        elif opcion == "2":
            # Cargar Vías
            cargaExcel("via", 11, 4)
            continue
        elif opcion == "3":
            # Cargar Autorizaciones
            cargaAutorizacion()
            continue
        elif opcion == "0":
            print("Saliendo del programa.")
            break
        else:
            print("Opción no válida. Intente de nuevo.\n")
            continue
if __name__ == "__main__":
 menuCargaExcel()
