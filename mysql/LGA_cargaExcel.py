import signal, sys
import os
from openpyxl import load_workbook
import re
import mysql.connector
import unicodedata
from mysql.connector import Error
from db.connection import get_connection


def cortarPrograma(sig, frame):
    print("\n\n[!] Saliendo ...\n")
    sys.exit(1)


signal.signal(signal.SIGINT, cortarPrograma)

# Rutas
script_dir = os.path.dirname(os.path.abspath(__file__))
docs_dir = os.path.join(script_dir, "docs")

fichero_permisos = os.path.join(docs_dir, "excel-cod.xlsx")

# Conexión MySQL
conn = get_connection()
cursor = conn.cursor()

wb = load_workbook(fichero_permisos)

# Normalizar validez

def normalizar_validez(valor):
    if not valor:
        return None, None

    v = str(valor).strip().lower()
    numeros = re.findall(r"\d+", v)

    if not numeros:
        return None, None

    num = int(numeros[0])

    if "mes" in v or v.endswith("m"):
        return num, "M"

    if "dia" in v or "días" in v or "dias" in v or v.isdigit():
        return num, "D"

    return None, None


def normalizar_sn(valor):
    if not valor:
        return "N"

    v = str(valor).strip().lower()
    return "N" if v == "n/a" else "S"


def calcular_dos_veces_smi(valor):
    if valor is None:
        return "N"

    texto = str(valor).strip()

    if texto == "":
        return "N"
    
    # Detecta 3.2 / 3.2.1 / 10.4.3 dentro del texto
    if re.search(r'\d+(?:\.\d+)+', texto):
        return "S"

    return "N"


def normalizar_autoriza_trabajar(valor):
    if not valor:
        return "N"

    # Convertir a string
    v = str(valor).strip().lower()

    # Quitar tildes
    v = ''.join(
        c for c in unicodedata.normalize('NFD', v)
        if unicodedata.category(c) != 'Mn'
    )

    # Casos NO claros
    if v in {"no", "/"}:
        return "N"

    # Si empieza por "si"
    if v.startswith("si"):
        return "S"

    # Casos laborales explícitos
    if "cuenta ajena" in v or "c/a" in v or "c/p" in v:
        return "S"

    # Por defecto
    return "N"



# Carga de datos

def cargaModelos():
    cursor.execute("DELETE FROM lga_modelos")

    for nombre_hoja in wb.sheetnames:
        try:
            cursor.execute(
                "INSERT INTO lga_modelos (ID, DES_MODELO) VALUES (%s, %s)",
                (nombre_hoja, "")
            )
        except Error as e:
            print(f"Error insertando modelo {nombre_hoja}: {e}")

    conn.commit()
    print("Modelos cargados")


def cargaPermisos():
    cursor.execute("DELETE FROM lga_permisos")

    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            id_permiso = fila[10]
            des_permiso = fila[3]
            lucrativo = normalizar_sn(fila[19])
            residencia = normalizar_sn(fila[18])
            via_defecto = fila[11]
            reglamento = fila[12]

            if not id_permiso or not des_permiso:
                continue

            try:
                cursor.execute(
                    """
                    INSERT INTO lga_permisos
                    (ID, DES_PERMISO, LUCRATIVO, RESIDENCIA, VIA_DEFECTO, REGLAMENTO)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    """,
                    (id_permiso, des_permiso, lucrativo, residencia, via_defecto, reglamento)
                )
            except Error as e:
                print(f"Error permiso fila {i}: {e}")

    conn.commit()
    print("Permisos cargados")


def cargaViaAcceso():
    cursor.execute("DELETE FROM lga_via_acceso")

    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            id_via = fila[11]
            des_via = fila[5]

            if not id_via:
                continue

            try:
                cursor.execute(
                    "INSERT INTO lga_via_acceso (ID, DES_VIA_ACCESO) VALUES (%s, %s)",
                    (id_via, des_via)
                )
            except Error as e:
                print(f"Error vía acceso fila {i}: {e}")

    conn.commit()
    print("Vías de acceso cargadas")


def cargaAutorizaciones():
    cursor.execute("DELETE FROM lga_autorizaciones")

    for nombre_hoja in wb.sheetnames:
        hoja = wb[nombre_hoja]
        id_modelo = nombre_hoja

        for i, fila in enumerate(hoja.iter_rows(min_row=2, values_only=True), start=2):
            cod_meyss = fila[9]
            id_permiso = fila[10]
            id_via = fila[11]
            silencio = fila[8]
            epigrafe_tasa_052 = fila[18]
            epigrafe_tasa_062 = fila[19]
            dos_veces_smi = calcular_dos_veces_smi(fila[20])
            autoriza_trabajar = normalizar_autoriza_trabajar(fila[21])

            num_plazo, tipo_plazo = normalizar_validez(fila[7])

            if not cod_meyss or not id_permiso or not id_via:
                continue

            try:
                cursor.execute(
                    """
                    INSERT INTO lga_autorizaciones
                    (COD_MEYSS, ID_PERMISO, ID_VIA, ID_MODELO,
                     NUM_PLAZO, TIPO_PLAZO, SILENCIO,
                     EPIGRAFE_TASA_052, EPIGRAFE_TASA_062, DOS_VECES_SMI, AUTORIZA_TRABAJAR)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        cod_meyss,
                        id_permiso,
                        id_via,
                        id_modelo,
                        num_plazo,
                        tipo_plazo,
                        silencio,
                        epigrafe_tasa_052,
                        epigrafe_tasa_062,
                        dos_veces_smi, 
                        autoriza_trabajar
                    )
                )
            except Error as e:
                print(f"Error autorización fila {i}: {e}")

    conn.commit()
    print("Autorizaciones cargadas correctamente")



# Menú

def menu():
    print("\nSeleccione una opción:")
    print("1. Cargar Modelos")
    print("2. Cargar Permisos")
    print("3. Cargar Vías de Acceso")
    print("4. Cargar Autorizaciones")
    print("0. Salir\n")


def menuExcel():
    while True:
        menu()
        opcion = input("Opción: ")

        if opcion == "1":
            cargaModelos()
        elif opcion == "2":
            cargaPermisos()
        elif opcion == "3":
            cargaViaAcceso()
        elif opcion == "4":
            cargaAutorizaciones()
        elif opcion == "0":
            cursor.close()
            conn.close()
            break
        else:
            print("Opción no válida")


if __name__ == "__main__":
    menuExcel()    
